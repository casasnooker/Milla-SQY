
# app_min_scrape_v14.py
# Tabs:
# 1) Réservations
# 2) Plan de dispatch
# 3) Plan de route (missions regroupées par navette + filtre statuts)
#
# Change vs v13:
# - Route tab groups by SHUTTLE (navette) instead of run.
# - Runs are still used internally to compute missions, but display is per shuttle, sorted by mission time.

ROUTE_SETTINGS_URLS = ["https://milla-sqy.millaapp.fr/api/v1/reservation/shuttle-route-settings/all"]

import os
import json
import time
import hashlib
import html
from datetime import date as dt_date, datetime
from typing import Any, Dict, List, Optional, Tuple

import requests

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from fastapi import FastAPI, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from dotenv import load_dotenv

load_dotenv()

# -----------------------------
# UI visibility switch (Python parameter)
# -----------------------------
# True  => show all tabs (admin / full UI)
# False => driver-only (only /mission, no visible tabs, other pages redirect)
CONFIG_SHOW_ALL_TABS: bool = True

# -----------------------------
# Driver-only helpers (authoritative)
# -----------------------------
def _iso_today() -> str:
    # ISO date YYYY-MM-DD based on local server time
    try:
        return str(dt_date.today())
    except Exception:
        import datetime as _dt
        return _dt.date.today().isoformat()

def _driver_only_mode() -> bool:
    # True => driver-only strict
    return not bool(CONFIG_SHOW_ALL_TABS)

TOKEN = os.getenv("MILLA_BEARER")

API_RESERVATIONS = "https://milla-sqy.millaapp.fr/api/v1/reservation/book/all-users"
API_STATIONS = "https://milla-sqy.millaapp.fr/api/v1/reservation/stations/all-duplicated"
API_TIME_ARRAY = "https://milla-sqy.millaapp.fr/api/v1/reservation/book/get-time-array"
API_SERVICE_SETTINGS_ALL = "https://milla-sqy.millaapp.fr/api/v1/reservation/service-settings/all"

API_ROUTE_SETTINGS_CANDIDATES = [
    "https://milla-sqy.millaapp.fr/api/v1/reservation/shuttle-route-settings/all",
    "https://milla-sqy.millaapp.fr/api/v1/reservation/route-settings/all",
    "https://milla-sqy.millaapp.fr/api/v1/reservation/shuttle-route-settings",
    "https://milla-sqy.millaapp.fr/api/v1/reservation/route-settings",
]

# "browser-like" headers
BROWSER_HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "fr-FR,fr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Connection": "keep-alive",
    "Content-Type": "application/json",
    "Origin": "https://milla-sqy.millaapp.fr",
    "Referer": "https://milla-sqy.millaapp.fr/service-settings",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-origin",
}

CONFIG: Dict[str, Any] = {
    # tolérance pour associer une réservation à un run en comparant heure resa vs heure théorique au stop de départ
    "route_plan_match_tolerance_minutes": 25,
    # si dep time est vide au stop, on essaie l'arrival time
    "route_plan_fallback_to_arrival_if_depart_missing": True,
    # mission repositionnement pré-service par navette (minutes avant le 1er départ)
    "pre_service_reposition_minutes": 15,

    # --- Auto-refresh web ---
    # Le front poll /state et recharge la page uniquement si la signature change.
    "web": {
        "poll_state_ms": 1000,
        "pause_refresh_ms_on_ui": 3000,

        # Cache côté serveur pour éviter de spammer l'API MILLA
        # (les pages et /state peuvent être rechargés souvent).
        "state_ttl_seconds": 1.0,
        "cache_resa_seconds": 1.0,
        "cache_time_array_seconds": 1.0,
        "cache_services_seconds": 60.0,
        "cache_stations_seconds": 300.0,
    },
}


# -----------------------------
# In-memory caching layer (avec TTL)
# -----------------------------
# Objectif: limiter les appels aux APIs (stations, services, time-array, reservations)
# tout en permettant l'auto-refresh (donc caches courts pour les données volatiles).
_CACHE: Dict[str, Dict[str, Any]] = {}

def _cache_get(key: str) -> Any:
    rec = _CACHE.get(key)
    if not rec:
        return None
    ttl = rec.get("ttl")
    ts = float(rec.get("ts", 0.0))
    if ttl is None:
        return rec.get("value")
    if (time.time() - ts) < float(ttl):
        return rec.get("value")
    # expiré
    _CACHE.pop(key, None)
    return None

def _cache_set(key: str, value: Any, ttl_seconds: Optional[float] = None) -> Any:
    _CACHE[key] = {"value": value, "ts": time.time(), "ttl": ttl_seconds}
    return value

def _cache_key(prefix: str, *parts: Any) -> str:
    return prefix + ":" + "|".join(str(p) for p in parts)

# -----------------------------
# Excel travel-time matrix loader

import unicodedata
import re

def _norm_excel(name: str) -> str:
    s = (name or "").strip().lower()
    s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    return s
# -----------------------------
# Reads the first .xlsx in the current folder (excluding temporary files),
# from sheet "temps" (fallback: first sheet).
#
# Matrix layout (as per your file):
# - Depart station names: column C, rows 4..34   (C4:C34)
# - Destination station names: row 3, columns D..AH (D3:AH3)
# - Travel minutes at intersection (row of depart, column of destination)
#
# Returned dict key: (norm(depart), norm(dest)) -> minutes (int)
_TRAVEL_MINUTES: Dict[Tuple[str, str], int] = {}
_ETA_DEBUG_MISS: int = 0
_ETA_DEBUG_MAX_MISS: int = 25


def _find_excel_in_cwd() -> Optional[str]:
    cwd = os.getcwd()
    for name in os.listdir(cwd):
        if not name.lower().endswith(".xlsx"):
            continue
        if name.startswith("~$"):
            continue
        return os.path.join(cwd, name)
    return None

def _ws_get_sheet(wb, preferred: str = "temps") -> Worksheet:
    if preferred in wb.sheetnames:
        return wb[preferred]
    # fallback: first sheet
    return wb[wb.sheetnames[0]]

def _to_int_minutes(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    try:
        if isinstance(v, (int, float)):
            if v != v:
                return None
            return int(round(float(v)))
        s = str(v).strip().replace(",", ".")
        if not s:
            return None
        return int(round(float(s)))
    except Exception:
        return None

def load_travel_minutes_from_excel() -> Tuple[Dict[Tuple[str, str], int], Optional[str]]:
    path = _find_excel_in_cwd()
    if not path:
        return {}, "Aucun fichier .xlsx trouvé dans le dossier courant (pour matrice temps)."
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return {}, f"Impossible d'ouvrir le fichier Excel '{os.path.basename(path)}': {e}"

    try:
        ws = _ws_get_sheet(wb, preferred="temps")
    except Exception as e:
        return {}, f"Feuille 'temps' introuvable et fallback impossible: {e}"

    # Read headers
    dep_names: List[Tuple[int, str]] = []
    for r in range(4, 35):  # 4..34
        name = ws[f"C{r}"].value
        if name is None or str(name).strip() == "":
            continue
        dep_names.append((r, str(name).strip()))

    dest_cols: List[Tuple[int, str]] = []
    # D..AH = 4..34
    for c in range(4, 35):
        val = ws.cell(row=3, column=c).value
        if val is None or str(val).strip() == "":
            continue
        dest_cols.append((c, str(val).strip()))

    table: Dict[Tuple[str, str], int] = {}
    for r, dep in dep_names:
        dep_n = _norm_excel(dep)
        for c, dst in dest_cols:
            dst_n = _norm_excel(dst)
            minutes = _to_int_minutes(ws.cell(row=r, column=c).value)
            if minutes is None:
                continue
            table[(dep_n, dst_n)] = minutes

    return table, None

def get_travel_minutes(dep: str, dst: str) -> Optional[int]:
    if not dep or not dst:
        return None
    dep_n = _norm_excel(dep)
    dst_n = _norm_excel(dst)
    return _TRAVEL_MINUTES.get((dep_n, dst_n))

def add_minutes_to_hhmmss(hhmmss: str, minutes: int) -> str:
    try:
        h, m, s = hhmmss.split(":")
        base = int(h) * 60 + int(m)
        total = base + int(minutes)
        total = total % (24 * 60)
        hh = total // 60
        mm = total % 60
        return f"{hh:02d}:{mm:02d}:00"
    except Exception:
        return ""



def _driver_only_redirect(day: str | None = None) -> RedirectResponse | None:
    """If driver-only mode enabled, return a RedirectResponse to /mission for the given day."""
    if bool(CONFIG_SHOW_ALL_TABS):
        return None
    d = day or _iso_today()
    return RedirectResponse(url=f"/mission?day={html.escape(str(d))}&kiosk=1")

app = FastAPI(title="Planification Missions - SQY Flex (Resa + Dispatch + Route)")

# Shuttle naming (UI)
SHUTTLE_ID_TO_NAME: Dict[int, str] = {
    61: "MB1",
    65: "MB5",
}


def sort_pins_natural(pins: List[str]) -> List[str]:
    """Sort passenger pins like P3, P12 in numeric order."""
    def _key(s: str) -> int:
        m = re.search(r"(\d+)", str(s))
        return int(m.group(1)) if m else 10**9
    try:
        return sorted([str(x) for x in (pins or [])], key=_key)
    except Exception:
        return [str(x) for x in (pins or [])]

def shuttle_label(shuttle_id: Any) -> str:
    try:
        sid = int(shuttle_id)
    except Exception:
        return f"Navette {shuttle_id}"
    name = SHUTTLE_ID_TO_NAME.get(sid)
    return f"Navette {name}" if name else f"Navette {sid}"

# Load travel-time matrix once (best effort)
try:
    _TRAVEL_MINUTES, _tm_err = load_travel_minutes_from_excel()
    if _tm_err:
        print("[WARN] Matrice Excel temps non chargée:", _tm_err)
    if not _TRAVEL_MINUTES:
        print("[WARN] Matrice Excel temps VIDE: ETA ne pourra pas être calculé (hors pré-service).")
    else:
        print(f"[INFO] Matrice Excel temps chargée: {len(_TRAVEL_MINUTES)} valeurs")
        # debug samples
        try:
            _sample = list(_TRAVEL_MINUTES.items())[:5]
            print("[DEBUG] Exemples matrice (dep,dst)->min:", _sample)
        except Exception:
            pass
except Exception as _e:
    print("[WARN] Erreur chargement matrice Excel temps:", _e)


def _auth_headers() -> Dict[str, str]:
    if not TOKEN:
        return {}
    return {"Authorization": f"Bearer {TOKEN}"}


def _headers(extra: Optional[Dict[str, str]] = None) -> Dict[str, str]:
    h: Dict[str, str] = {}
    h.update(BROWSER_HEADERS)
    h.update(_auth_headers())
    if extra:
        h.update(extra)
    return h


def _http_error_message(r: requests.Response) -> str:
    if r.status_code == 401:
        return "401 Unauthorized: token invalide/expiré (MILLA_BEARER)."
    if r.status_code == 403:
        return "403 Forbidden: accès refusé (droits/headers requis)."
    return f"HTTP {r.status_code}: {r.text[:500]}"


def _parse_iso_date(s: str) -> Optional[dt_date]:
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def _extract_hhmmss(s: Any) -> Optional[str]:
    if s is None:
        return None
    txt = str(s).strip()
    if not txt:
        return None
    if len(txt) >= 8 and txt[-8:].count(":") == 2:
        return txt[-8:]
    if len(txt) >= 5 and txt[-5:].count(":") == 1:
        return txt[-5:] + ":00"
    return None


def _hhmmss_to_minutes(hhmmss: Optional[str]) -> Optional[int]:
    if not hhmmss:
        return None
    try:
        h, m, s = hhmmss.split(":")
        return int(h) * 60 + int(m) + (1 if int(s) >= 30 else 0)
    except Exception:
        return None


def _normalize_station_name(name: Any) -> str:
    if name is None:
        return ""
    n = str(name).strip().lower()
    while "  " in n:
        n = n.replace("  ", " ")
    return n


def _normalize_station_name_loose(name: Any) -> str:
    n = _normalize_station_name(name)
    if n.endswith(" 1") or n.endswith(" 2"):
        n2 = n[:-2].strip()
        if n2:
            return n2
    return n



def assign_pins_by_shuttle(resa_rows: List[Dict[str, Any]]) -> None:
    """
    Add per-shuttle Passenger Identification Number (PIN) to each reservation row:
      - First pickup of the day for a given shuttle => P1
      - Second pickup => P2, etc.
    Numbering is independent per shuttle (shuttle 61 has its own P1, shuttle 65 has its own P1, ...).
    The result is stored in:
      - row["_pin"] : int
      - row["_pin_label"] : "P{_pin}"
    Sorting key (per shuttle): reservationTime (HH:MM:SS) then reservationId/execId as tie-breaker.
    """
    grouped: Dict[Any, List[Tuple[int, str, Dict[str, Any]]]] = {}
    for r in resa_rows:
        sh = r.get("shuttleId")
        if sh is None:
            continue
        hh = _extract_hhmmss(r.get("reservationTime"))
        tmin = _hhmmss_to_minutes(hh)
        tkey = tmin if tmin is not None else 10**9
        rid = r.get("reservationId")
        if rid is None:
            rid = r.get("execId")
        rid_str = str(rid) if rid is not None else ""
        grouped.setdefault(sh, []).append((tkey, rid_str, r))

    for sh, items in grouped.items():
        items.sort(key=lambda x: (x[0], x[1]))
        pin = 0
        for _, __, row in items:
            pin += 1
            row["_pin"] = pin
            row["_pin_label"] = f"P{pin}"



def _raw_fetch_reservations_for_day(day: str, size: int = 200, max_pages: int = 50) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    if not TOKEN:
        return [], "Token manquant: ajoute MILLA_BEARER dans le fichier .env"

    all_rows: List[Dict[str, Any]] = []
    page = 0
    while page < max_pages:
        params = {"page": page, "size": size, "start": day, "end": day}
        try:
            r = requests.get(API_RESERVATIONS, headers=_headers(), params=params, timeout=25)
        except Exception as e:
            return all_rows, f"Erreur réseau: {e}"

        if not r.ok:
            return all_rows, _http_error_message(r)

        try:
            payload = r.json()
        except Exception:
            return all_rows, f"Réponse non-JSON: {r.text[:500]}"

        data = payload.get("data", {})
        content = data.get("content", [])
        if not isinstance(content, list):
            return all_rows, f"Format inattendu: data.content est {type(content)}"

        all_rows.extend(content)

        last = bool(data.get("last", True))
        if last:
            break
        page += 1

    return all_rows, None




def fetch_reservations_for_day(day: str, size: int = 200, max_pages: int = 50) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    key = _cache_key("resa", day, size, max_pages)
    hit = _cache_get(key)
    if hit is not None:
        return hit
    out = _raw_fetch_reservations_for_day(day=day, size=size, max_pages=max_pages)
    ttl = float(CONFIG.get("web", {}).get("cache_resa_seconds", 1.0))
    return _cache_set(key, out, ttl_seconds=ttl)
def _raw_fetch_stations() -> Tuple[List[Dict[str, Any]], Optional[str]]:
    if not TOKEN:
        return [], "Token manquant: ajoute MILLA_BEARER dans le fichier .env"
    try:
        r = requests.get(API_STATIONS, headers=_headers(), timeout=25)
    except Exception as e:
        return [], f"Erreur réseau stations: {e}"
    if not r.ok:
        return [], f"Stations: {_http_error_message(r)}"
    try:
        payload = r.json()
    except Exception:
        return [], f"Stations: réponse non-JSON: {r.text[:500]}"
    data = payload.get("data", payload)
    if not isinstance(data, list):
        return [], f"Stations: format inattendu pour data: {type(data)}"
    return data, None




def fetch_stations() -> Tuple[List[Dict[str, Any]], Optional[str]]:
    key = _cache_key("stations")
    hit = _cache_get(key)
    if hit is not None:
        return hit
    out = _raw_fetch_stations()
    ttl = float(CONFIG.get("web", {}).get("cache_stations_seconds", 300.0))
    return _cache_set(key, out, ttl_seconds=ttl)
def _raw_fetch_service_settings_all() -> Tuple[List[Dict[str, Any]], Optional[str]]:
    if not TOKEN:
        return [], "Token manquant: ajoute MILLA_BEARER dans le fichier .env"
    try:
        r = requests.get(API_SERVICE_SETTINGS_ALL, headers=_headers(), timeout=25)
    except Exception as e:
        return [], f"Erreur réseau service-settings/all: {e}"
    if not r.ok:
        return [], f"Service-settings/all: {_http_error_message(r)}"
    try:
        payload = r.json()
    except Exception:
        return [], f"Service-settings/all: réponse non-JSON: {r.text[:500]}"
    data = payload.get("data", payload)
    if not isinstance(data, list):
        return [], f"Service-settings/all: format inattendu pour data: {type(data)}"
    return data, None





def _raw_fetch_route_settings_all() -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """Fetch route settings (ordered station_ids) from backend. Tries multiple endpoints."""
    if not TOKEN:
        return [], "Token manquant: ajoute MILLA_BEARER dans le fichier .env"
    last_err: Optional[str] = None
    for url in API_ROUTE_SETTINGS_CANDIDATES:
        try:
            r = requests.get(url, headers=_headers(), timeout=25)
        except Exception as e:
            last_err = f"Erreur réseau route-settings: {e}"
            continue
        if not r.ok:
            last_err = f"Route-settings: {_http_error_message(r)}"
            continue
        try:
            payload = r.json()
        except Exception:
            last_err = f"Route-settings: réponse non-JSON: {r.text[:500]}"
            continue
        data = payload.get("data", payload)
        if isinstance(data, list) and any(isinstance(x, dict) and "station_ids" in x for x in data):
            return data, None
        last_err = f"Route-settings: format inattendu data={type(data)}"
    return [], last_err or "Route-settings: aucune endpoint valide"

def fetch_route_settings_all() -> Tuple[List[Dict[str, Any]], Optional[str]]:
    key = _cache_key("route_settings_all")
    hit = _cache_get(key)
    if hit is not None:
        return hit
    out = _raw_fetch_route_settings_all()
    ttl = float(CONFIG.get("web", {}).get("cache_route_settings_seconds", 120.0))
    return _cache_set(key, out, ttl_seconds=ttl)

def _pick_best_route_setting(rows: List[Dict[str, Any]], shuttle_id: Optional[int], direction: str, n: int) -> Optional[Dict[str, Any]]:
    """Choose the best route setting for shuttle+direction. Prefer 'Parcours SQY' non-test, then length match."""
    if shuttle_id is None:
        return None
    diru = str(direction or "").upper()
    cands: List[Dict[str, Any]] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        if r.get("selectedShuttleId") != shuttle_id:
            continue
        if str(r.get("direction", "")).upper() != diru:
            continue
        sids = r.get("station_ids", [])
        if not isinstance(sids, list) or not sids:
            continue
        cands.append(r)
    if not cands:
        return None

    def score(r: Dict[str, Any]) -> Tuple[int, int, int, int]:
        desc = str(r.get("description", "")).lower()
        is_prod = 1 if ("parcours sqy" in desc and "test" not in desc and "démo" not in desc and "demo" not in desc) else 0
        sids = r.get("station_ids", [])
        ln = len(sids) if isinstance(sids, list) else 0
        # exact match preferred, then closest >= n, then max length
        exact = 1 if ln == n else 0
        ge = 1 if ln >= n else 0
        rid = int(r.get("id", 0)) if isinstance(r.get("id", 0), int) else 0
        return (exact, is_prod, ge * 1000 + min(ln, 999), rid)

    return sorted(cands, key=score, reverse=True)[0]

def fetch_service_settings_all() -> Tuple[List[Dict[str, Any]], Optional[str]]:
    key = _cache_key("services_all")
    hit = _cache_get(key)
    if hit is not None:
        return hit
    out = _raw_fetch_service_settings_all()
    ttl = float(CONFIG.get("web", {}).get("cache_services_seconds", 60.0))
    return _cache_set(key, out, ttl_seconds=ttl)
def select_service_for_day(services: List[Dict[str, Any]], day: str) -> Optional[Dict[str, Any]]:
    target = _parse_iso_date(day)
    if target is None:
        return None

    matches: List[Tuple[int, int, Dict[str, Any]]] = []
    for s in services:
        sd = _parse_iso_date(str(s.get("startDate", "")))
        ed = _parse_iso_date(str(s.get("endDate", "")))
        sid = s.get("id")
        if sd is None or ed is None or not isinstance(sid, int):
            continue
        if sd <= target <= ed:
            span_days = (ed - sd).days
            matches.append((span_days, -sid, s))

    if not matches:
        return None

    matches.sort(key=lambda x: (x[0], x[1]))
    return matches[0][2]


def _raw_fetch_time_array(service_id: int, service_start: str, service_end: str, route_ids: List[int]) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    if not TOKEN:
        return [], "Token manquant: ajoute MILLA_BEARER dans le fichier .env"

    url = f"{API_TIME_ARRAY}?serviceId={service_id}"
    body = {"serviceStart": service_start, "serviceEnd": service_end, "shuttleRouteIds": route_ids}

    try:
        r = requests.post(url, headers=_headers(), data=json.dumps(body), timeout=35)
    except Exception as e:
        return [], f"Erreur réseau time-array: {e}"

    if not r.ok:
        return [], f"Time-array: {_http_error_message(r)}"

    try:
        payload = r.json()
    except Exception:
        return [], f"Time-array: réponse non-JSON: {r.text[:500]}"

    data = payload.get("data", payload)
    if not isinstance(data, list):
        return [], f"Time-array: format inattendu pour data: {type(data)}"
    return data, None




def fetch_time_array(service_id: int, service_start: str, service_end: str, route_ids: List[int]) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    key = _cache_key("time_array", service_id, service_start, service_end, ",".join(str(x) for x in route_ids))
    hit = _cache_get(key)
    if hit is not None:
        return hit
    out = _raw_fetch_time_array(service_id=service_id, service_start=service_start, service_end=service_end, route_ids=route_ids)
    ttl = float(CONFIG.get("web", {}).get("cache_time_array_seconds", 1.0))
    return _cache_set(key, out, ttl_seconds=ttl)


# -----------------------------
# Signatures stables (pour auto-refresh)
# -----------------------------

def _sha1_bytes(b: bytes) -> str:
    return hashlib.sha1(b).hexdigest()


def _stable_reservations_signature(rows: List[Dict[str, Any]]) -> str:
    """Signature stable des réservations (suffisant pour détecter une évolution visible)."""
    slim: List[Tuple[Any, ...]] = []
    for r in rows or []:
        rid = r.get("reservationId")
        if rid is None:
            rid = r.get("execId")
        slim.append((
            rid,
            r.get("reservationTime"),
            r.get("status"),
            r.get("shuttleId"),
            r.get("routeId"),
            r.get("departStationName"),
            r.get("arriveStationName"),
        ))
    slim.sort(key=lambda x: (str(x[0]), str(x[1])))
    payload = json.dumps(slim, ensure_ascii=False, separators=(",", ":"))
    return _sha1_bytes(payload.encode("utf-8"))


def _stable_time_array_signature(items: List[Dict[str, Any]]) -> str:
    slim: List[Tuple[Any, ...]] = []
    for it in items or []:
        slim.append((
            it.get("shuttle"),
            it.get("direction"),
            it.get("time"),
            it.get("quota_vector"),
        ))
    slim.sort(key=lambda x: (str(x[0]), str(x[1])))
    payload = json.dumps(slim, ensure_ascii=False, separators=(",", ":"))
    return _sha1_bytes(payload.encode("utf-8"))


def _stable_route_signature(resa_kept: List[Dict[str, Any]], time_items: List[Dict[str, Any]], pre_minutes: int) -> str:
    base = _stable_reservations_signature(resa_kept) + "|" + _stable_time_array_signature(time_items) + f"|pre={int(pre_minutes)}"
    return _sha1_bytes(base.encode("utf-8"))


# Cache léger pour /state (poll toutes les ~1s)
_STATE_CACHE: Dict[str, Any] = {"key": None, "ts": 0.0, "sig": ""}


def compute_state_sig_cached(
    *,
    day: str,
    view: str,
    size: int,
    include_cancelled: int,
    statuses: Optional[str],
    statuses_locked: Optional[str],
    pre_minutes: Optional[int],
) -> Tuple[str, Optional[str]]:
    ttl = float(CONFIG.get("web", {}).get("state_ttl_seconds", 1.0))
    # Normalize statuses: empty string without explicit apply should behave like None (defaults)
    is_locked = str(statuses_locked or '').strip() == '1'
    norm_statuses = statuses
    if norm_statuses == '' and not is_locked:
        norm_statuses = None
    key = f"{view}|{day}|{size}|{int(include_cancelled or 0)}|{pre_minutes if pre_minutes is not None else ''}"
    now = time.time()
    if _STATE_CACHE.get("key") == key and (now - float(_STATE_CACHE.get("ts", 0.0))) < ttl:
        return str(_STATE_CACHE.get("sig") or ""), None

    # Reservations (used in resa + route)
    resa_rows: List[Dict[str, Any]] = []
    if view in {"resa", "route", "mission"}:
        resa_rows, err = fetch_reservations_for_day(day=day, size=size)
        if err:
            return "", err


        # Default: exclude CANCELLED unless include_cancelled=1
        if not bool(int(include_cancelled or 0)):
            resa_rows = [r for r in resa_rows if str(r.get("status", "") or "").strip() != "CANCELLED"]


    # Dispatch/time-array (used in dispatch + route)
    time_items: List[Dict[str, Any]] = []
    if view in {"dispatch", "route", "mission"}:
        stations, err_s = fetch_stations()
        if err_s:
            return "", err_s

        services, err_all = fetch_service_settings_all()
        if err_all:
            return "", err_all

        svc = select_service_for_day(services, day)
        if svc is None:
            return "", f"Aucun service trouvé pour day={day}"

        service_id = svc.get("id")
        service_start = str(svc.get("startTime", "") or "")
        service_end = str(svc.get("endTime", "") or "")
        route_ids = svc.get("shuttle_route_parameters_ids") or []

        if not isinstance(service_id, int):
            return "", f"Service id invalide pour day={day}"
        if not isinstance(route_ids, list) or not route_ids:
            return "", f"shuttle_route_parameters_ids manquant pour serviceId={service_id}"

        time_items, err_t = fetch_time_array(
            service_id=service_id,
            service_start=service_start,
            service_end=service_end,
            route_ids=[int(x) for x in route_ids],
        )
        if err_t:
            return "", err_t

        # 'stations' is only used to render; signature doesn't need it. Keep variable to avoid linter warnings.
        _ = stations

    if view == "dispatch":
        sig = _stable_time_array_signature(time_items)
    elif view in ("route","mission"):
        pre = int(pre_minutes) if pre_minutes is not None else int(CONFIG.get("pre_service_reposition_minutes", 15))
        sig = _stable_route_signature(resa_rows, time_items, pre)
    else:
        sig = _stable_reservations_signature(resa_rows)

    _STATE_CACHE.update({"key": key, "ts": now, "sig": sig})
    return sig, None
def build_station_sequence(stations: List[Dict[str, Any]], direction: str, n: int, shuttle_id: Optional[int] = None) -> List[Dict[str, Any]]:
    """Build the ordered station sequence for a run.

    Priority:
      1) Use Shuttle Route Settings (station_ids ordered) when available for this shuttle+direction.
      2) Fallback to legacy templates (if any) for known sizes.
      3) Fallback to displayIndex grouping.

    IMPORTANT: RETURN is NOT necessarily the inverse of FORWARD here.
    """
    # 1) Route settings (authoritative)
    rs_all, _err_rs = fetch_route_settings_all()
    if rs_all:
        picked = _pick_best_route_setting(rs_all, shuttle_id, direction, n)
        if picked:
            sids = picked.get("station_ids", [])
            id_map: Dict[int, Dict[str, Any]] = {}
            for s in stations:
                sid = s.get("id")
                if isinstance(sid, int):
                    id_map[sid] = s
            seq = []
            for sid in sids:
                try:
                    sid_int = int(sid)
                except Exception:
                    sid_int = None
                if sid_int is not None and sid_int in id_map:
                    seq.append(id_map[sid_int])
                else:
                    seq.append({"id": sid_int, "name": f"Station {sid_int}"})
            # fit to n
            if n <= len(seq):
                return seq[:n]
            seq.extend([{} for _ in range(n - len(seq))])
            return seq

    # 2) Legacy templates (kept for compatibility)
    by_name: Dict[str, Dict[str, Any]] = {str(s.get("name", "")).strip(): s for s in stations if s.get("name")}

    template_8 = [
        "Gare Routière des Prés 1",
        "SQY Ouest 2",
        "Fulgence Bienvenüe 2",
        "Les Chênes 2",
        "Pas du Lac",
        "Vieil Etang 1",
        "Vélodrome 1",
        "Gare Routière Paul Delouvrier 1",
    ]
    template_9 = [
        "Gare Routière Paul Delouvrier 1",
        "Vélodrome 2",
        "Vieil Etang 2",
        "Pas du Lac",
        "Les Quadrants 1",
        "Les Chênes 1",
        "Fulgence Bienvenüe 1",
        "SQY Ouest 1",
        "Gare Routière des Prés 1",
    ]

    if n == 8:
        return [by_name.get(name, {"name": name}) for name in template_8]
    if n == 9:
        return [by_name.get(name, {"name": name}) for name in template_9]

    # 3) Fallback: displayIndex grouping
    by_disp: Dict[int, List[Dict[str, Any]]] = {}
    for s in stations:
        if not s.get("active", True):
            continue
        di = s.get("displayIndex")
        if isinstance(di, int):
            by_disp.setdefault(di, []).append(s)

    def pick_variant(group: List[Dict[str, Any]]) -> Dict[str, Any]:
        if not group:
            return {}
        # Try to pick stable variant: prefer suffix based on direction heuristic
        if str(direction or "").upper() == "RETURN":
            for x in group:
                if str(x.get("name", "")).strip().endswith("1"):
                    return x
        for x in group:
            if str(x.get("name", "")).strip().endswith("2"):
                return x
        return group[0]

    disp_indices = sorted(by_disp.keys())
    seq: List[Dict[str, Any]] = [pick_variant(by_disp[di]) for di in disp_indices]

    if n <= len(seq):
        return seq[:n]
    seq.extend([{} for _ in range(n - len(seq))])
    return seq


def _html_shell(title: str, active: str, content: str, tip: str = "", current_sig: str = "", show_tabs: bool = True) -> str:
    # __SHELL_DRIVER_FORCE__
    if _driver_only_mode():
        show_tabs = False

    tabs = f"""
    <div class="tabs">
      <a class="tab {'active' if active=='resa' else ''}" href="/resa">Réservations</a>
      <a class="tab {'active' if active=='dispatch' else ''}" href="/dispatch">Plan de dispatch</a>
      <a class="tab {'active' if active=='route' else ''}" href="/route">Plan de route</a>
      <a class="tab {'active' if active=='mission' else ''}" href="/mission">Plan de mission</a>
    </div>
    """
    style = """
    <style>
      body { font-family: Arial, sans-serif; padding: 16px; background:#fff; }
      h1 { margin: 6px 0 14px 0; font-size: 20px; }
      .tabs { display:flex; gap:10px; align-items:center; margin: 6px 0 14px 0; }
      .tab { padding:10px 12px; border:1px solid #ddd; border-radius:12px; text-decoration:none; color:#111; background:#fafafa; }
      .tab.active { background:#111; color:#fff; border-color:#111; }
      .bar { display:flex; gap:12px; flex-wrap:wrap; align-items:flex-end; margin: 10px 0 14px 0; }
      .card { border:1px solid #eee; border-radius:14px; padding:10px 12px; background:#fcfcfc; }
      .card label { font-size: 12px; color:#555; display:block; margin-bottom:6px; }
      .card input { padding:8px 10px; border-radius:10px; border:1px solid #ddd; min-width: 170px; }
      .card button { padding:9px 12px; border-radius:12px; border:1px solid #111; background:#111; color:#fff; cursor:pointer; }
      table { border-collapse: collapse; width: 100%; margin-top: 10px; }
      th, td { border-bottom: 1px solid #eee; padding: 10px; text-align: left; vertical-align: top; }
      th { background: #f3f3f3; font-size: 13px; }
      td { font-size: 13px; }
      .err { color:#b00020; font-weight:700; margin-top: 12px; }
      .muted { color:#666; font-size: 12px; }
      .group { border:1px solid #eee; border-radius:16px; padding:12px; margin: 12px 0; background:#fff; }
      .group h3 { margin: 0 0 8px 0; font-size: 15px; }
      .pill { display:inline-block; padding:3px 10px; border-radius:999px; border:1px solid #ddd; background:#fafafa; font-size: 12px; color:#333; margin-left: 8px; }
      .tip { margin-top: 6px; }
      details { border:1px solid #eee; border-radius:14px; padding:10px 12px; background:#fff; }
      summary { cursor:pointer; font-weight:700; }
      .checks { display:flex; flex-wrap:wrap; gap:10px; margin-top:10px; }
      .chk { display:flex; gap:6px; align-items:center; border:1px solid #eee; border-radius:999px; padding:6px 10px; background:#fafafa; }
      .tiny { font-size:12px; font-weight:400; }
      .ok { color:#0a7f2e; font-weight:700; }
            /* Table alignment: keep all cells nicely aligned */
      th, td { vertical-align: middle; line-height: 1.25; }
      td div { line-height: 1.25; }

      /* ETA pill: keep icon + text on one line */
      .pill-eta { display:inline-flex; align-items:center; gap:6px; white-space:nowrap; }

      /* ETA INLINE LOCK */
      .pill-eta{ display:inline-flex; align-items:center; gap:6px; white-space:nowrap; flex-wrap:nowrap; }
      .pill-eta .ico{ display:inline-flex; align-items:center; line-height:1; }
      .pill-eta .ico svg{ width:16px; height:16px; display:block; }
      .pill-eta .txt{ display:inline; white-space:nowrap; }

      .pill-eta svg, .pill-eta span { display:inline-block; vertical-align:middle; }

      /* Objectifs: bold (driver + admin) */
      .objwrap, .objwrap * { font-weight: 700; }

      /* ETA: keep icon + text on one line (robust) */
      .pill-eta { display:inline-flex; align-items:center; gap:6px; white-space:nowrap; flex-wrap:nowrap; }
      .pill-eta > * { display:inline-block; vertical-align:middle; line-height:1; }
</style>
    """
    poll_ms = int(CONFIG.get("web", {}).get("poll_state_ms", 1000))
    pause_ms = int(CONFIG.get("web", {}).get("pause_refresh_ms_on_ui", 3000))

    # Auto-refresh: poll /state et reload si la signature change.
    # On pause temporairement si l'utilisateur interagit avec la page (input/checkbox/scroll/clavier).
    auto_refresh = f"""
    <script>
      let currentSig = {json.dumps(current_sig or "")};
      let pauseUntil = 0;
      const POLL_MS = {poll_ms};
      const PAUSE_MS = {pause_ms};

      function nowMs(){{ return Date.now(); }}
      function pauseRefresh(){{ pauseUntil = nowMs() + PAUSE_MS; }}

      ["mousedown","touchstart","focusin","keydown","wheel"].forEach(evt => {{
        document.addEventListener(evt, pauseRefresh, {{ passive:true }});
      }});

      async function pollState(){{
        try{{
          if(nowMs() < pauseUntil) return;
          const qs = window.location.search || "";
          const sep = qs.includes("?") ? "&" : "?";
          const url = "/state" + qs + sep + "view=" + encodeURIComponent({json.dumps(active)});
          const r = await fetch(url, {{ cache: "no-store" }});
          if(!r.ok) return;
          const data = await r.json();
          if(data && data.sig && data.sig !== currentSig){{
            window.location.reload();
          }}
        }} catch(e){{
          // silence
        }}
      }}
      setInterval(pollState, POLL_MS);
    

      // --- Mission tab day sync (set from Route tab, used by Mission tab) ---
      (function(){{
        try{{
          const key = "mission_day";
          // If a route page control exists, bind it
          const md = document.getElementById("missionDay");
          const btn = document.getElementById("openMission");
          function upd(val){{
            if(!val) return;
            localStorage.setItem(key, val);
            if(btn){{ btn.href = "/mission?day=" + encodeURIComponent(val); }}
          }}
          if(md){{
            md.addEventListener("change", ()=> upd(md.value));
            if(md.value){{ upd(md.value); }}
          }}
          // Intercept clicks on Mission tab to inject day
          document.querySelectorAll("a.tab").forEach(a=>{{
            if((a.getAttribute("href")||"") === "/mission"){{
              a.addEventListener("click", (e)=>{{
                const v = localStorage.getItem(key);
                if(v){{
                  a.setAttribute("href", "/mission?day=" + encodeURIComponent(v));
                }}
              }});
              // also set at load (so right-click/open new tab works)
              const v = localStorage.getItem(key);
              if(v){{
                a.setAttribute("href", "/mission?day=" + encodeURIComponent(v));
              }}
            }}
          }});
        }}catch(_){{}}
      }})();
    
</script>
    """

    return f"""<!doctype html>
<html><head><meta charset="utf-8"/><title>{title}</title>{style}</head>
<body>
{tabs if show_tabs else ''}
{content}
<div class="muted tip">{tip}</div>
{auto_refresh}
</body></html>"""


@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    r = _driver_only_redirect(request.query_params.get("day") or None)
    if r:
        return r
    return RedirectResponse(url="/resa")

@app.get("/state", response_class=JSONResponse)
def state(
    day: str = Query(default=str(dt_date.today()), description="YYYY-MM-DD"),
    view: str = Query(default="resa", description="resa, dispatch, route, mission"),
    size: int = Query(default=200, ge=1, le=200),
    include_cancelled: int = Query(default=0, ge=0, le=1, description="1 to include CANCELLED"),
    statuses: Optional[str] = Query(default=None, description="(deprecated)"),
    statuses_locked: Optional[str] = Query(default=None, description="(deprecated)"),
    pre_minutes: Optional[int] = Query(default=None, ge=0, le=180),
):
    """Endpoint léger pour l'auto-refresh (le front compare uniquement une signature)."""
    v = (view or "resa").strip().lower()
    if v not in {"resa", "dispatch", "route", "mission"}:
        v = "resa"

    sig, err = compute_state_sig_cached(day=day, view=v, size=size, include_cancelled=include_cancelled, statuses=statuses, statuses_locked=statuses_locked, pre_minutes=pre_minutes)
    if err:
        return {"ok": False, "day": day, "view": v, "error": err, "sig": "" , "server_ts": int(time.time())}
    return {"ok": True, "day": day, "view": v, "sig": sig, "server_ts": int(time.time())}


@app.get("/resa", response_class=HTMLResponse)
def page_resa(
    day: str = Query(default=str(dt_date.today()), description="YYYY-MM-DD"),
    include_cancelled: int = Query(default=0, ge=0, le=1, description="1 to include CANCELLED"),
    size: int = Query(default=50, ge=1, le=200),
):
    rows, err = fetch_reservations_for_day(day=day, size=size)

    bar = f"""
    <h1>Réservations</h1>
    <div class="bar">
      <div class="card">
        <label>day (YYYY-MM-DD)</label>
        <input id="day" value="{day}"/>
      </div>
      <div class="card">
        <label>size</label>
        <input id="size" value="{size}"/>
      </div>
      <div class="card">
        <label style="display:flex; gap:8px; align-items:center; cursor:pointer;">
          <input type="checkbox" id="add_cancelled" style="transform:scale(1.1);" {"checked" if int(include_cancelled or 0)==1 else ""}/>
          <span>Add cancelled</span>
        </label>
      </div>
      <div class="card">
        <button onclick="go()">Charger</button>
      </div>

      <div class="card">
        <div class="muted">Date onglet Mission</div>
        <input type="date" id="missionDay" value="{day}"/>
        <div style="height:6px"></div>
        <a id="openMission" class="tab" href="/mission?day={day}" style="display:inline-block; padding:8px 10px; border-radius:10px; border:1px solid #ddd; background:#fff; color:#111; text-decoration:none; font-size:13px;">Ouvrir</a>
      </div>

      <div class="card">
        <div class="muted">Total: <b>{len(rows)}</b></div>
      </div>
    </div>
    """
    bar += r"""
    <script>
      function go(){
        const d = document.getElementById('day').value;
        const s = document.getElementById('size').value;
        const addC = document.getElementById('add_cancelled') && document.getElementById('add_cancelled').checked ? '1' : '0';
        const u = new URL(window.location.href);
        const p = document.getElementById('pre') ? document.getElementById('pre').value : '';
        u.searchParams.set('day', d);
        u.searchParams.set('size', s);
        if(p !== '') u.searchParams.set('pre_minutes', p);
        u.searchParams.set('include_cancelled', addC);
        window.location.href = u.toString();
      }
    </script>
    """


    if err:
        return HTMLResponse(_html_shell("Réservations", "resa", bar + f"<div class='err'>❌ {err}</div>", current_sig=""))

    if not bool(int(include_cancelled or 0)):
        rows = [r for r in rows if str(r.get("status", "") or "").strip() != "CANCELLED"]

    # PIN numbering (per navette)
    assign_pins_by_shuttle(rows)

    def td(v: Any) -> str:
        return "" if v is None else str(v)

    head = """
      <tr>
        <th>Heure</th>
        <th>Statut</th>
        <th>Trajet</th>
        <th>Navette</th>
        <th>Départ</th>
        <th>Arrivée</th>
        <th>ID</th>
        <th>Exec</th>
        <th>PIN</th>
      </tr>
    """
    body = ""
    for r in rows:
        t = td(r.get("reservationTime", ""))
        status = td(r.get("status", ""))
        route_id = td(r.get("routeId", ""))
        shuttle = td(r.get("shuttleId", ""))
        dep = td(r.get("departStationName", ""))
        arr = td(r.get("arriveStationName", ""))
        rid = td(r.get("reservationId", ""))
        ex = td(r.get("execId", ""))
        pin = td(r.get("_pin_label", ""))
        body += f"<tr><td>{t}</td><td>{status}</td><td>{route_id}</td><td>{shuttle}</td><td><b>{dep}</b></td><td><b>{arr}</b></td><td>{rid}</td><td>{ex}</td><td><b>{pin}</b></td></tr>"

    table = f"<table><thead>{head}</thead><tbody>{body or '<tr><td colspan=9>Aucune réservation</td></tr>'}</tbody></table>"
    sig = _stable_reservations_signature(rows)
    return HTMLResponse(_html_shell("Réservations", "resa", bar + table, current_sig=sig))


@app.get("/dispatch", response_class=HTMLResponse)
def page_dispatch(
    day: str = Query(default=str(dt_date.today()), description="YYYY-MM-DD"),
):
    # Always show date selector first
    bar = f"""
    <h1>Plan de dispatch</h1>
    <div class="bar">
      <div class="card">
        <label>day (YYYY-MM-DD)</label>
        <input id="day" value="{day}"/>
      </div>
      <div class="card">
        <button onclick="go()">Charger</button>
      </div>
    </div>
    """
    bar += r"""
    <script>
      function go(){
        const d = document.getElementById('day').value;
        const u = new URL(window.location.href);
        u.searchParams.set('day', d);
        window.location.href = u.toString();
      }
    </script>
    """

    stations, err_s = fetch_stations()
    if err_s:
        return HTMLResponse(_html_shell("Plan de dispatch", "dispatch", bar + f"<div class='err'>❌ {err_s}</div>", current_sig=""))

    services, err_all = fetch_service_settings_all()
    if err_all:
        return HTMLResponse(_html_shell("Plan de dispatch", "dispatch", bar + f"<div class='err'>❌ {err_all}</div>", current_sig=""))

    svc = select_service_for_day(services, day)
    if svc is None:
        return HTMLResponse(_html_shell("Plan de dispatch", "dispatch", bar + f"<div class='err'>❌ Aucun service trouvé pour day={day}<br/><span class='muted'>Aucun service ne couvre cette date.</span></div>", current_sig=""))

    service_id = svc.get("id")
    service_start = str(svc.get("startTime", "") or "")
    service_end = str(svc.get("endTime", "") or "")
    route_ids = svc.get("shuttle_route_parameters_ids") or []

    if not isinstance(service_id, int):
        return HTMLResponse(_html_shell("Plan de dispatch", "dispatch", bar + f"<div class='err'>❌ Service id invalide pour day={day}</div>", current_sig=""))
    if not isinstance(route_ids, list) or not route_ids:
        return HTMLResponse(_html_shell("Plan de dispatch", "dispatch", bar + f"<div class='err'>❌ shuttle_route_parameters_ids manquant pour serviceId={service_id}</div>", current_sig=""))

    # Add small service summary card (optional)
    bar2 = bar + f"""
    <div class="bar">
      <div class="card">
        <label>serviceId</label>
        <input value="{service_id}" disabled/>
      </div>
      <div class="card">
        <label>serviceStart</label>
        <input value="{service_start}" disabled/>
      </div>
      <div class="card">
        <label>serviceEnd</label>
        <input value="{service_end}" disabled/>
      </div>
      <div class="card" style="min-width:260px;">
        <label>routeIds</label>
        <input value="{','.join(str(x) for x in route_ids)}" disabled style="min-width:240px;"/>
      </div>
    </div>
    """

    items, err_t = fetch_time_array(
        service_id=service_id,
        service_start=service_start,
        service_end=service_end,
        route_ids=[int(x) for x in route_ids],
    )
    if err_t:
        return HTMLResponse(_html_shell("Plan de dispatch", "dispatch", bar2 + f"<div class='err'>❌ {err_t}</div>", current_sig=""))

    groups_html = ""
    for idx, it in enumerate(items, start=1):
        shuttle = it.get("shuttle")
        direction = it.get("direction")
        time_arr = it.get("time", [])
        quota = it.get("quota_vector", [])
        n = len(time_arr) if isinstance(time_arr, list) else 0

        seq = build_station_sequence(stations, str(direction or ""), n, shuttle_id=int(shuttle) if isinstance(shuttle, int) else None)

        rows_html = ""
        for i in range(n):
            pair = time_arr[i] if i < len(time_arr) else [None, None]
            arr = pair[0] if isinstance(pair, list) and len(pair) > 0 else None
            dep = pair[1] if isinstance(pair, list) and len(pair) > 1 else None
            q = quota[i] if isinstance(quota, list) and i < len(quota) else ""
            st = seq[i] if i < len(seq) else {}
            name = st.get("name", f"Stop #{i+1}")
            di = st.get("displayIndex", "")
            dj = st.get("dispatchIndex", "")
            meta = f"<span class='muted'>dIdx:{dj} · disp:{di}</span>" if (di or dj) else "<span class='muted'>—</span>"
            rows_html += f"<tr><td><b>{name}</b><div>{meta}</div></td><td>{arr or ''}</td><td>{dep or ''}</td><td>{q}</td></tr>"

        table = f"""
        <table>
          <thead>
            <tr><th>Station</th><th>Arrivée</th><th>Départ</th><th>Quota</th></tr>
          </thead>
          <tbody>{rows_html or '<tr><td colspan=4>—</td></tr>'}</tbody>
        </table>
        """

        groups_html += f"""
        <div class="group">
          <h3>Run #{idx} <span class="pill">shuttle {shuttle}</span> <span class="pill">{direction}</span> <span class="pill">{n} stops</span></h3>
          {table}
        </div>
        """

    sig = _stable_time_array_signature(items)
    return HTMLResponse(_html_shell("Plan de dispatch", "dispatch", bar2 + groups_html, tip="Par défaut: today. Modifie day pour une autre date.", current_sig=sig))

def build_route_missions_by_shuttle(*, day: str, resa_rows: List[Dict[str, Any]], time_items: List[Dict[str, Any]], stations: List[Dict[str, Any]], pre_minutes: int) -> Dict[Any, List[Dict[str, Any]]]:
    """Construit les missions (regroupées par navette) à partir des réservations + time-array."""
    items = time_items
    pre = pre_minutes
    # Build runs model
    runs: List[Dict[str, Any]] = []
    for idx, it in enumerate(items, start=1):
        shuttle = it.get("shuttle")
        direction = it.get("direction")
        time_arr = it.get("time", [])
        if not isinstance(time_arr, list):
            continue
        n = len(time_arr)
        seq = build_station_sequence(stations, str(direction or ""), n, shuttle_id=int(shuttle) if isinstance(shuttle, int) else None)
        stops: List[Dict[str, Any]] = []
        for i in range(n):
            pair = time_arr[i] if i < len(time_arr) else [None, None]
            arr = pair[0] if isinstance(pair, list) and len(pair) > 0 else None
            dep = pair[1] if isinstance(pair, list) and len(pair) > 1 else None
            stx = seq[i] if i < len(seq) else {}
            name = stx.get("name", f"Stop #{i+1}")
            stops.append(
                {
                    "i": i,
                    "name": name,
                    "arr": _extract_hhmmss(arr),
                    "dep": _extract_hhmmss(dep),
                    "arr_min": _hhmmss_to_minutes(_extract_hhmmss(arr)),
                    "dep_min": _hhmmss_to_minutes(_extract_hhmmss(dep)),
                }
            )
        runs.append({"run_index": idx, "shuttle": shuttle, "direction": direction, "stops": stops})

    tol = int(CONFIG.get("route_plan_match_tolerance_minutes", 25))

    def stop_time_min(stop: Dict[str, Any]) -> Optional[int]:
        dm = stop.get("dep_min")
        am = stop.get("arr_min")
        if dm is not None:
            return dm
        if bool(CONFIG.get("route_plan_fallback_to_arrival_if_depart_missing", True)):
            return am
        return None

    def index_in_run(run: Dict[str, Any], station_name: str) -> Optional[int]:
        nn = _normalize_station_name(station_name)
        nl = _normalize_station_name_loose(station_name)
        for stp in run["stops"]:
            sn = _normalize_station_name(stp.get("name"))
            if sn == nn or sn == nl or _normalize_station_name_loose(sn) == nl:
                return int(stp["i"])
        return None

    activity_pick: Dict[int, Dict[int, int]] = {}
    activity_drop: Dict[int, Dict[int, int]] = {}
    pick_pins: Dict[int, Dict[int, List[str]]] = {}
    drop_pins: Dict[int, Dict[int, List[str]]] = {}

    unmatched = 0
    for r in resa_rows:
        shuttle_id = r.get("shuttleId")
        dep_name = str(r.get("departStationName", "") or "")
        arr_name = str(r.get("arriveStationName", "") or "")
        t_resa = _extract_hhmmss(r.get("reservationTime"))
        tmin = _hhmmss_to_minutes(t_resa)

        if shuttle_id is None or tmin is None or not dep_name or not arr_name:
            unmatched += 1
            continue

        best_run = None
        best_score = None

        for run in runs:
            if run.get("shuttle") != shuttle_id:
                continue
            i_dep = index_in_run(run, dep_name)
            i_arr = index_in_run(run, arr_name)
            if i_dep is None or i_arr is None or i_dep >= i_arr:
                continue
            tm_dep = stop_time_min(run["stops"][i_dep])
            if tm_dep is None:
                continue
            score = abs(tm_dep - tmin)
            if best_score is None or score < best_score:
                best_score = score
                best_run = run

        if best_run is None or best_score is None or best_score > tol:
            unmatched += 1
            continue

        ridx = int(best_run["run_index"])
        i_dep = index_in_run(best_run, dep_name)
        i_arr = index_in_run(best_run, arr_name)
        if i_dep is None or i_arr is None:
            unmatched += 1
            continue

        activity_pick.setdefault(ridx, {})
        activity_drop.setdefault(ridx, {})
        pick_pins.setdefault(ridx, {})
        drop_pins.setdefault(ridx, {})

        activity_pick[ridx][int(i_dep)] = activity_pick[ridx].get(int(i_dep), 0) + 1
        activity_drop[ridx][int(i_arr)] = activity_drop[ridx].get(int(i_arr), 0) + 1

        pin_label = str(r.get("_pin_label", "") or "")
        if pin_label:
            pick_pins[ridx].setdefault(int(i_dep), []).append(pin_label)
            drop_pins[ridx].setdefault(int(i_arr), []).append(pin_label)

    missions_by_shuttle: Dict[Any, List[Dict[str, Any]]] = {}
    total_missions = 0

    for run in runs:
        ridx = int(run["run_index"])
        shuttle = run.get("shuttle")
        direction = run.get("direction")

        pick_map = activity_pick.get(ridx, {})
        drop_map = activity_drop.get(ridx, {})

        nstops = len(run["stops"])
        if nstops == 0:
            continue

        active_stops = set(pick_map.keys()) | set(drop_map.keys())
        active_stops.add(0)
        active_stops.add(nstops - 1)

        active_sorted = sorted(active_stops)

        for order_idx, (a, b) in enumerate(zip(active_sorted[:-1], active_sorted[1:]), start=1):
            st_a = run["stops"][a]
            st_b = run["stops"][b]

            heure = st_a.get("dep") or (st_a.get("arr") if bool(CONFIG.get("route_plan_fallback_to_arrival_if_depart_missing", True)) else None)
            heure_min = _hhmmss_to_minutes(heure) if heure else None

            frm = st_a.get("name", f"Stop #{a+1}")
            to = st_b.get("name", f"Stop #{b+1}")

            pu = int(pick_map.get(b, 0))
            do = int(drop_map.get(b, 0))
            pu_pins = pick_pins.get(ridx, {}).get(b, [])
            do_pins = drop_pins.get(ridx, {}).get(b, [])

            m = {
                "heure": heure or "",
                "heure_min": heure_min if heure_min is not None else 10**9,
                "from": frm,
                "to": to,
                "pickup": pu,
                "dropoff": do,
                "pickup_pins": pu_pins,
                "dropoff_pins": do_pins,
                "run_index": ridx,
                "direction": direction,
                "order": order_idx,
            }

            missions_by_shuttle.setdefault(shuttle, []).append(m)
            total_missions += 1


    # --- Pre-service reposition mission per shuttle (always) ---
    # Creates one mission per shuttle, scheduled 'pre' minutes before the first terminus departure,
    # ending at the terminus of the earliest run. This also surfaces pick-ups at the terminus (stop 0).
    earliest_by_shuttle: Dict[Any, Dict[str, Any]] = {}
    for run in runs:
        shuttle = run.get("shuttle")
        stops = run.get("stops", [])
        if not stops:
            continue
        st0 = stops[0]
        t0 = st0.get("dep") or (st0.get("arr") if bool(CONFIG.get("route_plan_fallback_to_arrival_if_depart_missing", True)) else None)
        t0m = _hhmmss_to_minutes(t0) if t0 else None
        if t0m is None:
            continue
        cur = earliest_by_shuttle.get(shuttle)
        if cur is None or t0m < int(cur["t0m"]):
            earliest_by_shuttle[shuttle] = {
                "t0m": int(t0m),
                "terminus": st0.get("name", "Terminus"),
                "run_index": int(run.get("run_index", 0)),
                "direction": run.get("direction"),
            }

    # Ensure pre is non-negative and sensible
    pre = max(0, int(pre))

    for shuttle, meta in earliest_by_shuttle.items():
        t0m = int(meta["t0m"])
        pre_start_m = max(0, t0m - pre)  # clamp (no wrap to previous day)
        hh = pre_start_m // 60
        mm = pre_start_m % 60
        pre_hhmm = f"{hh:02d}:{mm:02d}:00"

        ridx0 = int(meta["run_index"])
        pu0 = int(activity_pick.get(ridx0, {}).get(0, 0))
        do0 = int(activity_drop.get(ridx0, {}).get(0, 0))
        pu_pins0 = pick_pins.get(ridx0, {}).get(0, [])
        do_pins0 = drop_pins.get(ridx0, {}).get(0, [])

        m = {
            "heure": pre_hhmm,
            "heure_min": pre_start_m,
            "from": "Repositionnement",
            "to": meta["terminus"],
            "pickup": pu0,
            "dropoff": do0,
            "pickup_pins": pu_pins0,
            "dropoff_pins": do_pins0,
            "run_index": ridx0,
            "direction": meta.get("direction"),
            "order": -1,
            "is_pre_service": True,
        }
        missions_by_shuttle.setdefault(shuttle, []).append(m)

    for sh, ms in missions_by_shuttle.items():
        ms.sort(key=lambda x: (x.get("heure_min", 10**9), x.get("run_index", 0), x.get("order", 0)))

    return missions_by_shuttle

@app.get("/route", response_class=HTMLResponse)
def page_route(
    day: str = Query(default=str(dt_date.today()), description="YYYY-MM-DD"),
    include_cancelled: int = Query(default=0, ge=0, le=1, description="1 to include CANCELLED"),
    statuses: Optional[str] = Query(default=None, description="(deprecated)"),
    statuses_locked: Optional[str] = Query(default=None, description="(deprecated)"),
    size: int = Query(default=200, ge=1, le=200),
    pre_minutes: Optional[int] = Query(default=None, ge=0, le=180, description="Minutes avant 1er départ pour mission pré-service"),
):
    # pre-service reposition minutes (resolve display value)
    pre = int(pre_minutes) if pre_minutes is not None else int(CONFIG.get("pre_service_reposition_minutes", 15))
    pre_minutes_display = str(pre)

    # Base bar always visible (even if errors later)
    bar_base = f"""
    <h1>Plan de route</h1>
    <div class="bar">
      <div class="card">
        <label>day (YYYY-MM-DD)</label>
        <input id="day" value="{day}"/>
      </div>
      <div class="card">
        <label>size (reservations/page)</label>
        <input id="size" value="{size}"/>
      </div>
      <div class="card">
        <label>Pré-service (min)</label>
        <input id="pre" value="{pre_minutes_display}"/>
      </div>
      <div class="card">
        <label style="display:flex; gap:8px; align-items:center; cursor:pointer;">
          <input type="checkbox" id="add_cancelled" style="transform:scale(1.1);" {"checked" if int(include_cancelled or 0)==1 else ""}/>
          <span>Add cancelled</span>
        </label>
      </div>
      <div class="card">
        <button onclick="applyAll()">Charger</button>
      </div>
    </div>
    """

    # --- reservations ---
    resa_rows, err_r = fetch_reservations_for_day(day=day, size=size)
    if err_r:
        bar_err = bar_base + r"""
        <script>
          function applyAll(){
            const d = document.getElementById('day').value;
            const s = document.getElementById('size').value;
            const u = new URL(window.location.href);
            u.searchParams.set('day', d);
            u.searchParams.set('size', s);
            window.location.href = u.toString();
          }
        </script>
        """
        return HTMLResponse(_html_shell("Plan de route", "route", bar_err + f"<div class='err'>❌ {err_r}</div>", current_sig=""))

    # PIN numbering (per navette) computed on ALL reservations of the day (stable even when filtering statuses)
    assign_pins_by_shuttle(resa_rows)

    all_statuses = sorted({str(r.get("status", "") or "").strip() for r in resa_rows if str(r.get("status", "") or "").strip()})

    inc_cancelled = bool(int(include_cancelled or 0))
    if inc_cancelled:
        selected_set = set(all_statuses)
    else:
        selected_set = set([s for s in all_statuses if s != "CANCELLED"])

    resa_kept = [r for r in resa_rows if str(r.get("status", "") or "").strip() in selected_set]

    # Reuse base bar (with Add cancelled checkbox) and add summary line
    bar = bar_base + f"""
    <div class=\"bar\">
      <div class=\"card\">
        <div class=\"muted\">Réservations gardées: <b>{len(resa_kept)}</b> / {len(resa_rows)}</div>
      </div>
    </div>
    """

    bar += r"""
    <script>
      function applyAll(){
        const d = document.getElementById('day').value;
        const s = document.getElementById('size').value;
        const p = document.getElementById('pre') ? document.getElementById('pre').value : '';
        const addC = document.getElementById('add_cancelled') && document.getElementById('add_cancelled').checked ? '1' : '0';

        const u = new URL(window.location.href);
        u.searchParams.set('day', d);
        u.searchParams.set('size', s);
        if(p !== '') u.searchParams.set('pre_minutes', p);
        u.searchParams.set('include_cancelled', addC);
        u.searchParams.delete('statuses');
        u.searchParams.delete('statuses_locked');
        window.location.href = u.toString();
      }
    </script>
    """


    # --- dispatch runs ---
    stations, err_s = fetch_stations()
    if err_s:
        return HTMLResponse(_html_shell("Plan de route", "route", bar + f"<div class='err'>❌ {err_s}</div>", current_sig=""))

    services, err_all = fetch_service_settings_all()
    if err_all:
        return HTMLResponse(_html_shell("Plan de route", "route", bar + f"<div class='err'>❌ {err_all}</div>", current_sig=""))

    svc = select_service_for_day(services, day)
    if svc is None:
        return HTMLResponse(_html_shell("Plan de route", "route", bar + f"<div class='err'>❌ Aucun service trouvé pour day={day}<br/><span class='muted'>Aucun service ne couvre cette date.</span></div>", current_sig=""))

    service_id = svc.get("id")
    service_start = str(svc.get("startTime", "") or "")
    service_end = str(svc.get("endTime", "") or "")
    route_ids = svc.get("shuttle_route_parameters_ids") or []

    if not isinstance(service_id, int):
        return HTMLResponse(_html_shell("Plan de route", "route", bar + f"<div class='err'>❌ Service id invalide pour day={day}</div>", current_sig=""))
    if not isinstance(route_ids, list) or not route_ids:
        return HTMLResponse(_html_shell("Plan de route", "route", bar + f"<div class='err'>❌ shuttle_route_parameters_ids manquant pour serviceId={service_id}</div>", current_sig=""))

    items, err_t = fetch_time_array(
        service_id=service_id,
        service_start=service_start,
        service_end=service_end,
        route_ids=[int(x) for x in route_ids],
    )
    if err_t:
        return HTMLResponse(_html_shell("Plan de route", "route", bar + f"<div class='err'>❌ {err_t}</div>", current_sig=""))

    # Build runs model
    runs: List[Dict[str, Any]] = []
    for idx, it in enumerate(items, start=1):
        shuttle = it.get("shuttle")
        direction = it.get("direction")
        time_arr = it.get("time", [])
        if not isinstance(time_arr, list):
            continue
        n = len(time_arr)
        seq = build_station_sequence(stations, str(direction or ""), n, shuttle_id=int(shuttle) if isinstance(shuttle, int) else None)
        stops: List[Dict[str, Any]] = []
        for i in range(n):
            pair = time_arr[i] if i < len(time_arr) else [None, None]
            arr = pair[0] if isinstance(pair, list) and len(pair) > 0 else None
            dep = pair[1] if isinstance(pair, list) and len(pair) > 1 else None
            stx = seq[i] if i < len(seq) else {}
            name = stx.get("name", f"Stop #{i+1}")
            stops.append(
                {
                    "i": i,
                    "name": name,
                    "arr": _extract_hhmmss(arr),
                    "dep": _extract_hhmmss(dep),
                    "arr_min": _hhmmss_to_minutes(_extract_hhmmss(arr)),
                    "dep_min": _hhmmss_to_minutes(_extract_hhmmss(dep)),
                }
            )
        runs.append({"run_index": idx, "shuttle": shuttle, "direction": direction, "stops": stops})

    tol = int(CONFIG.get("route_plan_match_tolerance_minutes", 25))

    def stop_time_min(stop: Dict[str, Any]) -> Optional[int]:
        dm = stop.get("dep_min")
        am = stop.get("arr_min")
        if dm is not None:
            return dm
        if bool(CONFIG.get("route_plan_fallback_to_arrival_if_depart_missing", True)):
            return am
        return None

    def index_in_run(run: Dict[str, Any], station_name: str) -> Optional[int]:
        nn = _normalize_station_name(station_name)
        nl = _normalize_station_name_loose(station_name)
        for stp in run["stops"]:
            sn = _normalize_station_name(stp.get("name"))
            if sn == nn or sn == nl or _normalize_station_name_loose(sn) == nl:
                return int(stp["i"])
        return None

    activity_pick: Dict[int, Dict[int, int]] = {}
    activity_drop: Dict[int, Dict[int, int]] = {}
    pick_pins: Dict[int, Dict[int, List[str]]] = {}
    drop_pins: Dict[int, Dict[int, List[str]]] = {}

    unmatched = 0
    for r in resa_kept:
        shuttle_id = r.get("shuttleId")
        dep_name = str(r.get("departStationName", "") or "")
        arr_name = str(r.get("arriveStationName", "") or "")
        t_resa = _extract_hhmmss(r.get("reservationTime"))
        tmin = _hhmmss_to_minutes(t_resa)

        if shuttle_id is None or tmin is None or not dep_name or not arr_name:
            unmatched += 1
            continue

        best_run = None
        best_score = None

        for run in runs:
            if run.get("shuttle") != shuttle_id:
                continue
            i_dep = index_in_run(run, dep_name)
            i_arr = index_in_run(run, arr_name)
            if i_dep is None or i_arr is None or i_dep >= i_arr:
                continue
            tm_dep = stop_time_min(run["stops"][i_dep])
            if tm_dep is None:
                continue
            score = abs(tm_dep - tmin)
            if best_score is None or score < best_score:
                best_score = score
                best_run = run

        if best_run is None or best_score is None or best_score > tol:
            unmatched += 1
            continue

        ridx = int(best_run["run_index"])
        i_dep = index_in_run(best_run, dep_name)
        i_arr = index_in_run(best_run, arr_name)
        if i_dep is None or i_arr is None:
            unmatched += 1
            continue

        activity_pick.setdefault(ridx, {})
        activity_drop.setdefault(ridx, {})
        pick_pins.setdefault(ridx, {})
        drop_pins.setdefault(ridx, {})

        activity_pick[ridx][int(i_dep)] = activity_pick[ridx].get(int(i_dep), 0) + 1
        activity_drop[ridx][int(i_arr)] = activity_drop[ridx].get(int(i_arr), 0) + 1

        pin_label = str(r.get("_pin_label", "") or "")
        if pin_label:
            pick_pins[ridx].setdefault(int(i_dep), []).append(pin_label)
            drop_pins[ridx].setdefault(int(i_arr), []).append(pin_label)

    missions_by_shuttle: Dict[Any, List[Dict[str, Any]]] = {}
    total_missions = 0

    for run in runs:
        ridx = int(run["run_index"])
        shuttle = run.get("shuttle")
        direction = run.get("direction")

        pick_map = activity_pick.get(ridx, {})
        drop_map = activity_drop.get(ridx, {})

        nstops = len(run["stops"])
        if nstops == 0:
            continue

        active_stops = set(pick_map.keys()) | set(drop_map.keys())
        active_stops.add(0)
        active_stops.add(nstops - 1)

        active_sorted = sorted(active_stops)

        for order_idx, (a, b) in enumerate(zip(active_sorted[:-1], active_sorted[1:]), start=1):
            st_a = run["stops"][a]
            st_b = run["stops"][b]

            heure = st_a.get("dep") or (st_a.get("arr") if bool(CONFIG.get("route_plan_fallback_to_arrival_if_depart_missing", True)) else None)
            heure_min = _hhmmss_to_minutes(heure) if heure else None

            frm = st_a.get("name", f"Stop #{a+1}")
            to = st_b.get("name", f"Stop #{b+1}")

            pu = int(pick_map.get(b, 0))
            do = int(drop_map.get(b, 0))
            pu_pins = pick_pins.get(ridx, {}).get(b, [])
            do_pins = drop_pins.get(ridx, {}).get(b, [])

            m = {
                "heure": heure or "",
                "heure_min": heure_min if heure_min is not None else 10**9,
                "from": frm,
                "to": to,
                "pickup": pu,
                "dropoff": do,
                "pickup_pins": pu_pins,
                "dropoff_pins": do_pins,
                "run_index": ridx,
                "direction": direction,
                "order": order_idx,
            }

            missions_by_shuttle.setdefault(shuttle, []).append(m)
            total_missions += 1


    # --- Pre-service reposition mission per shuttle (always) ---
    # Creates one mission per shuttle, scheduled 'pre' minutes before the first terminus departure,
    # ending at the terminus of the earliest run. This also surfaces pick-ups at the terminus (stop 0).
    earliest_by_shuttle: Dict[Any, Dict[str, Any]] = {}
    for run in runs:
        shuttle = run.get("shuttle")
        stops = run.get("stops", [])
        if not stops:
            continue
        st0 = stops[0]
        t0 = st0.get("dep") or (st0.get("arr") if bool(CONFIG.get("route_plan_fallback_to_arrival_if_depart_missing", True)) else None)
        t0m = _hhmmss_to_minutes(t0) if t0 else None
        if t0m is None:
            continue
        cur = earliest_by_shuttle.get(shuttle)
        if cur is None or t0m < int(cur["t0m"]):
            earliest_by_shuttle[shuttle] = {
                "t0m": int(t0m),
                "terminus": st0.get("name", "Terminus"),
                "run_index": int(run.get("run_index", 0)),
                "direction": run.get("direction"),
            }

    # Ensure pre is non-negative and sensible
    pre = max(0, int(pre))

    for shuttle, meta in earliest_by_shuttle.items():
        t0m = int(meta["t0m"])
        pre_start_m = max(0, t0m - pre)  # clamp (no wrap to previous day)
        hh = pre_start_m // 60
        mm = pre_start_m % 60
        pre_hhmm = f"{hh:02d}:{mm:02d}:00"

        ridx0 = int(meta["run_index"])
        pu0 = int(activity_pick.get(ridx0, {}).get(0, 0))
        do0 = int(activity_drop.get(ridx0, {}).get(0, 0))
        pu_pins0 = pick_pins.get(ridx0, {}).get(0, [])
        do_pins0 = drop_pins.get(ridx0, {}).get(0, [])

        m = {
            "heure": pre_hhmm,
            "heure_min": pre_start_m,
            "from": "Repositionnement",
            "to": meta["terminus"],
            "pickup": pu0,
            "dropoff": do0,
            "pickup_pins": pu_pins0,
            "dropoff_pins": do_pins0,
            "run_index": ridx0,
            "direction": meta.get("direction"),
            "order": -1,
            "is_pre_service": True,
        }
        missions_by_shuttle.setdefault(shuttle, []).append(m)

    for sh, ms in missions_by_shuttle.items():
        ms.sort(key=lambda x: (x.get("heure_min", 10**9), x.get("run_index", 0), x.get("order", 0)))

    content = bar + f"<div class='muted'>Missions générées: <b>{total_missions}</b> · Non matchées: <b>{unmatched}</b></div>"

    if not missions_by_shuttle:
        content += "<div class='muted'>Aucune mission.</div>"
    else:
        for shuttle, ms in sorted(missions_by_shuttle.items(), key=lambda kv: str(kv[0])):
            rows_html = ""
            for m in ms:
                pu = int(m.get("pickup", 0))
                do = int(m.get("dropoff", 0))

                act = "<span class='muted'>Repositionnement</span>"
                if pu or do:
                    parts = []
                    if pu:
                        pins = ', '.join(sort_pins_natural(m.get('pickup_pins', []) or []))
                        tail = f" <span class='tiny'>({pins})</span>" if pins else ""
                        parts.append(f"<span class='ok'>Pick up: {pu}</span>{tail}")
                    if do:
                        pins = ', '.join(sort_pins_natural(m.get('dropoff_pins', []) or []))
                        tail = f" <span class='tiny'>({pins})</span>" if pins else ""
                        parts.append(f"<span class='ok'>Drop off: {do}</span>{tail}")
                    act = " · ".join(parts)

                
                
                # ETA (toutes missions) = heure début + T(depart -> destination) depuis matrice Excel
                # IMPORTANT: la mission "pré-service" a un traitement à part -> pas d'ETA (position de départ inconnue)
                eta_txt = ""
                start_h = str(m.get('heure','') or '').strip()
                dep_eta = str(m.get('from','') or '').strip()
                dst_eta = str(m.get('to','') or '').strip()

                if start_h and (not bool(m.get('is_pre_service', False))):
                    minutes = get_travel_minutes(dep_eta, dst_eta)
                    if minutes is not None:
                        eta_txt = add_minutes_to_hhmmss(start_h, int(minutes))
                    else:
                        global _ETA_DEBUG_MISS
                        if _ETA_DEBUG_MISS < _ETA_DEBUG_MAX_MISS:
                            _ETA_DEBUG_MISS += 1
                            print(f"[DEBUG][ETA] manquant: dep='{dep_eta}' -> dst='{dst_eta}' | heure='{start_h}'")

                if eta_txt:
                    act = act + f"<div class='muted tiny' style='margin-top:4px;'>ETA: <b>{eta_txt}</b></div>"



                meta = f"<span class='muted'>run #{m.get('run_index')} · {m.get('direction')}</span>"
                rows_html += f"<tr><td><b>{m.get('heure','')}</b></td><td><b>{m.get('from','')}</b><div>{meta}</div></td><td><b>{m.get('to','')}</b></td><td>{act}</td></tr>"

            content += f"""
            <div class="group">
              <h3>Navette {shuttle} <span class="pill">{len(ms)} missions</span></h3>
              <table>
                <thead>
                  <tr><th>Heure mission</th><th>Départ</th><th>Destination</th><th>Activité (au stop destination)</th></tr>
                </thead>
                <tbody>
                  {rows_html or "<tr><td colspan=4 class='muted'>Aucune mission</td></tr>"}
                </tbody>
              </table>
            </div>
            """

    tip = f"Par défaut: today. Modifie day pour une autre date. Tolérance match={tol} min."
    sig = _stable_route_signature(resa_kept, items, pre)
    return HTMLResponse(_html_shell("Plan de route", "route", content, tip=tip, current_sig=sig))

@app.get("/mission", response_class=HTMLResponse)
def page_mission(
    shuttle: str = Query(default="ALL", description="ALL | MB1 | MB5"),
    day: str | None = Query(default=None, description="YYYY-MM-DD (optionnel). Par défaut: today"),
):
    """
    Onglet premium : même scraping + même génération que /route,
    UI simplifiée (date uniquement), CANCELLED exclu par défaut.
    """
    size = 200
    pre = int(CONFIG.get("pre_service_reposition_minutes", 15))

    day = str(dt_date.today()) if _driver_only_mode() else str(day or dt_date.today())

    # --- reservations ---
    resa_rows, err_r = fetch_reservations_for_day(day=day, size=size)
    if err_r:
        content = f"<h1>Plan de mission - SQY Flex</h1><div class='err'>{html.escape(str(err_r))}</div>"
        return HTMLResponse(_html_shell("Plan de mission", "mission", content, tip="", current_sig=""))

    # Exclure CANCELLED ici (pas de checkbox)
    resa_kept = [r for r in resa_rows if str(r.get("status", "") or "").strip() != "CANCELLED"]

    # --- dispatch/time-array ---
    stations, err_s = fetch_stations()
    if err_s:
        content = f"<h1>Plan de mission - SQY Flex</h1><div class='err'>{html.escape(str(err_s))}</div>"
        return HTMLResponse(_html_shell("Plan de mission", "mission", content, tip="", current_sig=""))

    services, err_all = fetch_service_settings_all()
    if err_all:
        content = f"<h1>Plan de mission - SQY Flex</h1><div class='err'>{html.escape(str(err_all))}</div>"
        return HTMLResponse(_html_shell("Plan de mission", "mission", content, tip="", current_sig=""))

    svc = select_service_for_day(services, day)
    if svc is None:
        content = f"<h1>Plan de mission - SQY Flex</h1><div class='err'>Aucun service trouvé pour day={html.escape(day)}</div>"
        return HTMLResponse(_html_shell("Plan de mission", "mission", content, tip="", current_sig=""))

    service_id = svc.get("id")
    service_start = str(svc.get("startTime", "") or "")
    service_end = str(svc.get("endTime", "") or "")
    route_ids = svc.get("shuttle_route_parameters_ids") or []
    if not isinstance(service_id, int):
        content = f"<h1>Plan de mission - SQY Flex</h1><div class='err'>Service id invalide pour day={html.escape(day)}</div>"
        return HTMLResponse(_html_shell("Plan de mission", "mission", content, tip="", current_sig=""))
    if not isinstance(route_ids, list) or not route_ids:
        content = f"<h1>Plan de mission - SQY Flex</h1><div class='err'>shuttle_route_parameters_ids manquant pour serviceId={service_id}</div>"
        return HTMLResponse(_html_shell("Plan de mission", "mission", content, tip="", current_sig=""))

    items, err_t = fetch_time_array(
        service_id=service_id,
        service_start=service_start,
        service_end=service_end,
        route_ids=[int(x) for x in route_ids],
    )
    if err_t:
        content = f"<h1>Plan de mission - SQY Flex</h1><div class='err'>{html.escape(str(err_t))}</div>"
        return HTMLResponse(_html_shell("Plan de mission", "mission", content, tip="", current_sig=""))

    # PIN numbering stable
    assign_pins_by_shuttle(resa_rows)

    missions_by_shuttle = build_route_missions_by_shuttle(
        day=day,
        resa_rows=resa_kept,
        time_items=items,
        stations=stations,
        pre_minutes=pre,
    )

    # Compte réservations par navette (pour l’affichage "N Missions (M Réservations)")
    def _norm_sh(x: Any) -> Any:
        try:
            return int(x)
        except Exception:
            return str(x)

    resa_count_by_shuttle: Dict[Any, int] = {}
    for r in resa_kept:
        sh = _norm_sh(r.get("shuttleId"))
        resa_count_by_shuttle[sh] = resa_count_by_shuttle.get(sh, 0) + 1

    # Optional shuttle filter (UI)
    sh_filter = (shuttle or "ALL").strip().upper()
    allowed_ids: Optional[set] = None
    if sh_filter in ("MB1", "MB5"):
        # Map name -> id (inverse of SHUTTLE_ID_TO_NAME)
        inv = {v.upper(): k for k, v in SHUTTLE_ID_TO_NAME.items()}
        sid = inv.get(sh_filter)
        if sid is not None:
            allowed_ids = {sid}

    if allowed_ids is not None:
        missions_by_shuttle = {k: v for k, v in missions_by_shuttle.items() if int(k) in allowed_ids}
        resa_count_by_shuttle = {k: v for k, v in resa_count_by_shuttle.items() if int(k) in allowed_ids}


    fancy_css = r"""
    <style>
      body{ background:#f4f6fb; }
      .hero{ display:flex; align-items:flex-start; justify-content:space-between; gap:18px; margin: 10px 0 18px 0;}
      .hero .left h1{ font-size:34px; margin:0; letter-spacing:-0.5px;}
      .hero .left .sub{ margin-top:6px; color:#667085; font-size:14px;}

      .cardx{ background:#fff; border:1px solid rgba(16,24,40,.08); box-shadow:0 10px 30px rgba(16,24,40,.08); border-radius:18px; padding:16px 18px; }
      .shuttle{ margin-top:14px; }
      .shuttle .head{ display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; }
      .shuttle .name{ font-size:18px; font-weight:900; }
      .summary-pill{ border-radius:999px; padding:8px 12px; border:1px solid rgba(16,24,40,.12); font-weight:900; background:#fff; }

      table{ width:100%; border-collapse:separate; border-spacing:0; overflow:hidden; border-radius:14px; }
      thead th{ background:#0f172a; color:#fff; font-size:13px; padding:12px 14px; border-bottom:0; }
      tbody td{ padding:14px; border-bottom:1px solid rgba(16,24,40,.08); font-size:13.5px; background:#fff; }
      tbody tr:last-child td{ border-bottom:0; }
      .col-heure{ width:120px; font-weight:900; }
      .dest strong{ font-weight:900; }
      .objwrap{ display:flex; gap:10px; align-items:center; flex-wrap:wrap; }

      .pill{ display:inline-flex; align-items:center; gap:8px; padding:8px 12px; border-radius:999px; border:1px solid rgba(2,6,23,.12); font-weight:700; line-height:1; }
      .pill svg{ width:15px; height:15px; display:block; }
      .pill .ico{ width:15px; height:15px; display:inline-flex; align-items:center; justify-content:center; text-align:center; font-size:15px; line-height:15px; position:relative; top:-1px; }
      .pill-pick{ border-color: rgba(34,197,94,.35); background: rgba(34,197,94,.12); }
      .pill-drop{ border-color: rgba(239,68,68,.35); background: rgba(239,68,68,.12); }
      .pill-repos{ border-color: rgba(148,163,184,.55); background: rgba(148,163,184,.20); }
      .pill-eta{ border-color: rgba(59,130,246,.35); background: rgba(59,130,246,.12); }

      .tiny{ font-size:12px; font-weight:800; }
      .muted2{ color:#667085; font-size:12px; }

      .hero .right{ display:flex; flex-direction:column; align-items:flex-end; gap:10px; }
      .controls{ display:flex; gap:10px; align-items:center; }
      .controls .date{ padding:10px 12px; border-radius:12px; border:1px solid #e4e7ec; background:#fff; font-weight:800; }
      .controls .btn{ padding:10px 14px; border-radius:14px; border:1px solid #111827; background:#111827; color:#fff; font-weight:900; cursor:pointer; }
      .controls .btn:hover{ filter:brightness(.95); }
    </style>
    """

    icon_clock = r"""<svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
      <path d="M12 7v5l3 2" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
      <path d="M12 22c5.523 0 10-4.477 10-10S17.523 2 12 2 2 6.477 2 12s4.477 10 10 10Z" stroke="currentColor" stroke-width="2"/>
    </svg>"""

    icon_pick = "<span class='ico'>↑</span>"

    icon_drop = "<span class='ico'>↓</span>"

    icon_repos = "<span class='ico'>→</span>"

    hero = f"""{fancy_css}
    <div class='hero'>
      <div class='left'>
        <h1>Plan de mission - SQY Flex</h1>
        <div class='sub'>Date de service : <b>{html.escape(day)}</b></div>
      </div>
      <div class='right'>
        <form class='controls' action='/mission' method='get'>
          <select class='date' name='shuttle'>
            <option value='ALL' {'selected' if (shuttle or 'ALL').strip().upper()=='ALL' else ''}>Toutes les navettes</option>
            <option value='MB1' {'selected' if (shuttle or '').strip().upper()=='MB1' else ''}>Navette MB1</option>
            <option value='MB5' {'selected' if (shuttle or '').strip().upper()=='MB5' else ''}>Navette MB5</option>
          </select>
          <button class='btn' type='submit'>Afficher</button>
        </form>
      </div>
    </div>
    """

    if not missions_by_shuttle:
        sig = _stable_route_signature(resa_kept, items, pre)
        content = hero + "<div class='cardx'><div class='muted2'>Aucune mission générée.</div></div>"
        return HTMLResponse(_html_shell("Plan de mission", "mission", content, tip="", current_sig=sig))

    parts: List[str] = [hero]

    for shuttle, ms in sorted(missions_by_shuttle.items(), key=lambda kv: str(kv[0])):
        sh_norm = _norm_sh(shuttle)
        ms_sorted = sorted(ms, key=lambda x: int(x.get('heure_min', 10**9)))

        rows = []
        for m in ms_sorted:
            hhmm = html.escape(str(m.get('heure','') or '').strip())

            dep = str(m.get('from','') or '').strip()
            if bool(m.get('is_pre_service', False)):
                dep_disp = "Position actuelle"
            else:
                dep_disp = dep
            dep_html = html.escape(dep_disp)

            dst = str(m.get('to','') or '').strip()
            dst_html = html.escape(dst)

            pu = int(m.get('pickup', 0) or 0)
            do = int(m.get('dropoff', 0) or 0)

            pills: List[str] = []

            # Objectif
            if pu == 0 and do == 0:
                pills.append(f"<span class='pill pill-repos'>{icon_repos} Repositionnement</span>")
            else:
                if pu:
                    pins = ', '.join(sort_pins_natural(m.get('pickup_pins', []) or []))
                    tail = f" <span class='tiny'>({html.escape(pins)})</span>" if pins else ""
                    pills.append(f"<span class='pill pill-pick'>{icon_pick} Pick up : {pu}{tail}</span>")
                if do:
                    pins = ', '.join(sort_pins_natural(m.get('dropoff_pins', []) or []))
                    tail = f" <span class='tiny'>({html.escape(pins)})</span>" if pins else ""
                    pills.append(f"<span class='pill pill-drop'>{icon_drop} Drop off : {do}{tail}</span>")

            # ETA (comme /route)
            eta_txt = ""
            start_h = str(m.get('heure','') or '').strip()
            dep_eta = str(m.get('from','') or '').strip()
            dst_eta = str(m.get('to','') or '').strip()
            if start_h and (not bool(m.get('is_pre_service', False))):
                minutes = get_travel_minutes(dep_eta, dst_eta)
                if minutes is not None:
                    eta_txt = add_minutes_to_hhmmss(start_h, int(minutes))

            if eta_txt:
                pills.append(f"<span class='pill pill-eta'>{icon_clock} ETA : {html.escape(eta_txt)}</span>")

            rows.append(f"""
              <tr>
                <td class='col-heure'>{hhmm}</td>
                <td>{dep_html}</td>
                <td class='dest'><strong>{dst_html}</strong></td>
                <td><div class='objwrap'>{''.join(pills)}</div></td>
              </tr>
            """)

        resa_count = int(resa_count_by_shuttle.get(sh_norm, 0))
        parts.append(f"""
        <div class='shuttle cardx'>
          <div class='head'>
            <div class='name'>{html.escape(shuttle_label(shuttle))}</div>
            <div class='summary-pill'>{len(ms_sorted)} Missions ({resa_count} Réservations)</div>
          </div>
          <table>
            <thead>
              <tr><th>Heure</th><th>Départ</th><th>Destination</th><th>Objectifs</th></tr>
            </thead>
            <tbody>{''.join(rows)}</tbody>
          </table>
        </div>
        """)

    content = "\n".join(parts)
    sig = _stable_route_signature(resa_kept, items, pre)
    return HTMLResponse(_html_shell("Plan de mission", "mission", content, tip="", current_sig=sig))



def _run():
    port = int(os.getenv("PORT", "8000"))
    try:
        import uvicorn  # type: ignore
    except Exception:
        print("uvicorn manquant. Installe requirements.txt puis relance.")
        raise
    uvicorn.run(app, host="127.0.0.1", port=port, log_level="info")


if __name__ == "__main__":
    _run()
